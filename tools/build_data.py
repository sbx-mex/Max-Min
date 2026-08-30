#!/usr/bin/env python3
"""Construye el motor compacto Max&Min desde CSV semanales y cruces Excel."""

from __future__ import annotations

import argparse
import csv
import io
import json
import math
import re
import shutil
import struct
import tempfile
import unicodedata
import zipfile
from collections import Counter, OrderedDict, defaultdict
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

RECORD = struct.Struct("<BHHI")
FILES_PER_FOLDER = 80


def text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def norm(value: object) -> str:
    value = unicodedata.normalize("NFD", text(value))
    value = "".join(char for char in value if unicodedata.category(char) != "Mn")
    value = re.sub(r"\s+", " ", value).strip().lower()
    return re.sub(r"[^a-z0-9]+", "", value)


def norm_variants(value: object) -> list[str]:
    base = norm(value)
    variants = [base]
    if base.startswith("zzz"):
        variants.append(base[3:])
    variants.append(re.sub(r"^0+", "", base))
    return list(dict.fromkeys(item for item in variants if item))


def rows_from_xlsx(path: Path, sheet_name: str) -> list[tuple[object, ...]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name]
    rows = list(worksheet.iter_rows(values_only=True))
    workbook.close()
    return rows


def normalized_csv_lines(binary: io.BufferedReader) -> list[str]:
    raw = binary.read()
    encoding = "utf-16" if raw.startswith((b"\xff\xfe", b"\xfe\xff")) else "utf-8-sig"
    lines = raw.decode(encoding).splitlines()
    output: list[str] = []
    for line in lines:
        value = line.strip()
        if not value or value.startswith("Max & Min_"):
            continue
        if '""' in value:
            if value.startswith('"') and value.endswith('"'):
                value = value[1:-1]
            value = value.replace('""', '"')
        output.append(value)
    return output


def parse_week(filename: str) -> int:
    match = re.search(r"_(\d+)\.csv$", filename, re.IGNORECASE)
    if not match:
        raise ValueError(f"No se pudo identificar la semana en {filename}")
    week = int(match.group(1))
    if not 1 <= week <= 53:
        raise ValueError(f"Semana fuera de rango en {filename}")
    return week


class SpoolWriter:
    def __init__(self, root: Path, limit: int = 96) -> None:
        self.root = root
        self.limit = limit
        self.handles: OrderedDict[str, io.BufferedWriter] = OrderedDict()

    def write(self, ceco: str, payload: bytes) -> None:
        handle = self.handles.pop(ceco, None)
        if handle is None:
            if len(self.handles) >= self.limit:
                _, oldest = self.handles.popitem(last=False)
                oldest.close()
            handle = (self.root / f"{ceco}.bin").open("ab")
        handle.write(payload)
        self.handles[ceco] = handle

    def close(self) -> None:
        for handle in self.handles.values():
            handle.close()
        self.handles.clear()


def load_crosses(directory_path: Path, price_path: Path, presentation_path: Path) -> tuple[dict[str, str], dict[str, tuple[str, str, str]], dict[str, dict[str, object]]]:
    directory_rows = rows_from_xlsx(directory_path, "Directorio")
    directory = {text(row[0]): text(row[1]) for row in directory_rows[1:] if row and text(row[0])}

    sap_rows = rows_from_xlsx(price_path, "SAP")
    sap_by_code = {
        text(row[1]): (text(row[0]), text(row[1]), text(row[2]))
        for row in sap_rows[1:]
        if row and len(row) >= 3 and text(row[1])
    }
    micros_rows = rows_from_xlsx(price_path, "Catalogo Micros")
    micros_by_name: dict[str, tuple[str, str, str]] = {}
    for row in micros_rows[1:]:
        if not row or len(row) < 2 or not text(row[1]):
            continue
        code = text(row[0])
        match = sap_by_code.get(code, ("", code, ""))
        for key in norm_variants(row[1]):
            current = micros_by_name.get(key)
            if current is None or (not current[2] and match[2]):
                micros_by_name[key] = match

    presentations_raw = json.loads(presentation_path.read_text(encoding="utf-8"))
    presentations = {norm(key): value for key, value in presentations_raw.items()}
    return directory, micros_by_name, presentations


def minified_json(value: object) -> str:
    return json.dumps(value, ensure_ascii=False, separators=(",", ":"))


def build(args: argparse.Namespace) -> dict[str, object]:
    output = args.output.resolve()
    data_root = output / "data"
    for old in data_root.glob("stores_*"):
        if old.is_dir():
            shutil.rmtree(old)
    data_root.mkdir(parents=True, exist_ok=True)
    (output / "audit").mkdir(parents=True, exist_ok=True)

    directory, micros_by_name, presentations = load_crosses(args.directory, args.prices, args.presentations)
    categories: list[str] = []
    category_ids: dict[str, int] = {}
    ingredients: list[dict[str, object]] = []
    ingredient_ids: dict[str, int] = {}
    ingredient_norm_ids: dict[str, int] = {}
    weeks: list[int] = []
    rows_by_week: dict[str, int] = {}
    positive_by_week: dict[str, int] = {}
    stores_with_data: set[str] = set()
    indicator_nonblank = 0
    raw_rows = 0

    with tempfile.TemporaryDirectory(prefix="maxmin-build-") as temp_name:
        spool_root = Path(temp_name)
        spool = SpoolWriter(spool_root)
        with zipfile.ZipFile(args.zip) as archive:
            names = sorted((name for name in archive.namelist() if name.lower().endswith(".csv")), key=parse_week)
            actual_weeks = [parse_week(name) for name in names]
            if not actual_weeks:
                raise ValueError("El ZIP no contiene CSV semanales")
            expected_end = args.expected_end or max(actual_weeks)
            expected_weeks = list(range(args.expected_start, expected_end + 1))
            if actual_weeks != expected_weeks:
                raise ValueError(f"Se esperaban semanas {args.expected_start}-{expected_end}; se recibieron {actual_weeks}")
            for name in names:
                week = parse_week(name)
                weeks.append(week)
                week_rows = 0
                week_positive = 0
                reader = csv.DictReader(normalized_csv_lines(archive.open(name)))
                required = {"Semana", "Tiendas", "Categoría Inventario", "Ingrediente", "Indicadores", "Uso Ideal* (#)"}
                if not required.issubset(set(reader.fieldnames or [])):
                    raise ValueError(f"Encabezados inválidos en {name}: {reader.fieldnames}")
                for row in reader:
                    week_rows += 1
                    raw_rows += 1
                    ceco = text(row.get("Tiendas"))
                    category = text(row.get("Categoría Inventario"))
                    source_name = text(row.get("Ingrediente"))
                    if text(row.get("Indicadores")):
                        indicator_nonblank += 1
                    try:
                        usage = float(text(row.get("Uso Ideal* (#)")).replace(",", "") or 0)
                    except ValueError:
                        usage = 0
                    if usage <= 0 or not ceco or not category or not source_name:
                        continue
                    if ceco not in directory:
                        raise ValueError(f"CeCo {ceco} de {name} no existe en Directorio.xlsx")
                    if category not in category_ids:
                        category_ids[category] = len(categories)
                        categories.append(category)
                    source_key = norm(source_name)
                    ingredient_id = ingredient_norm_ids.get(source_key)
                    if ingredient_id is None:
                        ingredient_id = len(ingredients)
                        ingredient_norm_ids[source_key] = ingredient_id
                        ingredient_ids[source_name] = ingredient_id
                        sap = next((micros_by_name[key] for key in norm_variants(source_name) if key in micros_by_name), ("", "", ""))
                        presentation = presentations.get(source_key, {})
                        unit = text(presentation.get("unidad")) or "PZA:"
                        pickpack = text(presentation.get("pickpack")) or unit
                        factor = max(1, int(float(presentation.get("factor") or 1)))
                        ingredients.append({
                            "name": source_name,
                            "sap": sap[2] or source_name,
                            "code": sap[1],
                            "woe": sap[0],
                            "unit": unit,
                            "pickpack": pickpack,
                            "factor": factor,
                            "sapStatus": "ok" if sap[2] else "review",
                            "formatStatus": "ok" if presentation else "review",
                        })
                    cents = max(0, min(4_294_967_295, int(round(usage * 100))))
                    spool.write(ceco, RECORD.pack(week, category_ids[category], ingredient_id, cents))
                    stores_with_data.add(ceco)
                    week_positive += 1
                rows_by_week[str(week)] = week_rows
                positive_by_week[str(week)] = week_positive
        spool.close()

        store_entries: list[dict[str, object]] = []
        sorted_codes = sorted(directory, key=lambda code: (int(code) if code.isdigit() else math.inf, code))
        data_codes = [code for code in sorted_codes if code in stores_with_data]
        for index, ceco in enumerate(data_codes):
            folder_number = index // FILES_PER_FOLDER + 1
            folder = data_root / f"stores_{folder_number:02d}"
            folder.mkdir(parents=True, exist_ok=True)
            target = folder / f"{ceco}.json"
            by_week: dict[str, list[int]] = defaultdict(list)
            raw = (spool_root / f"{ceco}.bin").read_bytes()
            if len(raw) % RECORD.size:
                raise ValueError(f"Spool dañado para {ceco}")
            for offset in range(0, len(raw), RECORD.size):
                week, category_id, ingredient_id, cents = RECORD.unpack_from(raw, offset)
                by_week[str(week)].extend((category_id, ingredient_id, cents))
            target.write_text(minified_json(dict(by_week)), encoding="utf-8")
            store_entries.append({
                "code": ceco,
                "name": directory[ceco],
                "label": f"{ceco} · {directory[ceco]}",
                "file": target.relative_to(output).as_posix(),
            })

    sap_ok = sum(item["sapStatus"] == "ok" for item in ingredients)
    format_ok = sum(item["formatStatus"] == "ok" for item in ingredients)
    manifest = {
        "version": "4.1-weekly",
        "generated": date.today().isoformat(),
        "weeks": weeks,
        "categories": categories,
        "ingredients": ingredients,
        "stores": store_entries,
        "storesWithoutData": [
            {"code": code, "name": directory[code]}
            for code in sorted(directory)
            if code not in stores_with_data
        ],
        "counts": {
            "directoryStores": len(directory),
            "storesWithData": len(stores_with_data),
            "categories": len(categories),
            "ingredients": len(ingredients),
            "rows": raw_rows,
            "positiveRows": sum(positive_by_week.values()),
            "rowsByWeek": rows_by_week,
            "positiveByWeek": positive_by_week,
            "sapMatched": sap_ok,
            "formatMatched": format_ok,
            "indicatorNonBlank": indicator_nonblank,
        },
        "formula": "Promedio de uso de semanas seleccionadas / 7. Máximo: 2 pedidos x5, 3 x4, 4 x3, 5 x2.",
    }
    (data_root / "manifest.js").write_text("window.MAXMIN_MANIFEST=" + minified_json(manifest) + ";\n", encoding="utf-8")
    report = {
        "status": "ok",
        "generated": manifest["generated"],
        "counts": manifest["counts"],
        "folders": {
            folder.name: {"files": len(list(folder.iterdir())), "bytes": sum(path.stat().st_size for path in folder.iterdir())}
            for folder in sorted(data_root.glob("stores_*"))
        },
    }
    (output / "audit" / "data_build_report.json").write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")

    review_path = output / "audit" / "ingredients_review.csv"
    with review_path.open("w", encoding="utf-8-sig", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=["Ingrediente", "Descripción SAP", "Código DIA", "ID WOE", "Cruce SAP", "Cruce Formato"])
        writer.writeheader()
        for item in ingredients:
            if item["sapStatus"] != "ok" or item["formatStatus"] != "ok":
                writer.writerow({
                    "Ingrediente": item["name"],
                    "Descripción SAP": item["sap"],
                    "Código DIA": item["code"],
                    "ID WOE": item["woe"],
                    "Cruce SAP": item["sapStatus"],
                    "Cruce Formato": item["formatStatus"],
                })
    return report


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--zip", type=Path, required=True)
    parser.add_argument("--directory", type=Path, required=True)
    parser.add_argument("--prices", type=Path, required=True)
    parser.add_argument("--presentations", type=Path, default=Path(__file__).with_name("presentation_reference.json"))
    parser.add_argument("--output", type=Path, default=Path(__file__).resolve().parents[1])
    parser.add_argument("--expected-start", type=int, default=1)
    parser.add_argument("--expected-end", type=int)
    return parser.parse_args()


if __name__ == "__main__":
    result = build(parse_args())
    print(json.dumps(result, ensure_ascii=False, indent=2))
